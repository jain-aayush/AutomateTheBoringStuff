"""
Write a program that performs the tasks of the previous program in reverse order:
The program should open a spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.
"""

import openpyxl

spreadsheet = input("Enter the name of the spreadsheet you want to open\n")

wb = openpyxl.load_workbook(filename = spreadsheet)
sheet = wb.active

max_cols = sheet.max_column
max_row = sheet.max_row

for col in range(1, max_cols+1):
    column_content = []                                     #storing the contents of a row in a list
    for row in range(1, max_row+1):
        value = sheet.cell(column=col, row=row).value
        column_content.append(value)
    filename = "col"+str(col)+".txt"                        #filename according to the column number
    with open(filename, 'w') as file:
        for item in column_content:
            if item is not None:                            #writing values which are not None to text file
                file.write("%s\n" % item)

