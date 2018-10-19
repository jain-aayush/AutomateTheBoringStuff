"""
Write a program to invert the row and column of the cells in the spreadsheet. 
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa). 
This should be done for all cells in the spreadsheet.
"""

import openpyxl

spreadsheet = input("Enter the name of the file to be inverted")

wb = openpyxl.load_workbook(filename = spreadsheet)
sheet = wb.active
wb_new = openpyxl.Workbook()
new_sheet = wb_new.active

max_cols = sheet.max_column
max_rows = sheet.max_row

for row in range(1, max_rows+1):
    for col in range(1, max_cols+1):
        new_sheet.cell(row = col, column= row).value = sheet.cell(row = row, column= col).value

wb_new.save(spreadsheet)