"""
Create a program multiplicationTable.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet.
"""

import openpyxl, sys

n = int(sys.argv[1])                                                    #converting the CLA into integer

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication Table"

fontObj = openpyxl.styles.Font(bold = True)                             #font object to make the cell content bold

for i in range(2,n+2):                                                  #making the table outline
    cell1 = sheet.cell(row = 1, column= i, value = i-1)
    cell1.font = fontObj
    cell2 = sheet.cell(row = i, column= 1, value = i-1)
    cell2.font = fontObj

for col in range(1,n+1):                                                #filling in the n*n matrix
    for row in range(1,n+1):
        cell = sheet.cell(row = row+1, column= col+1, value = row*col)

wb.save("Multiplication Table.xlsx")